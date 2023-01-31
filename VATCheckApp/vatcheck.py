import openpyxl
from zeep import Client
from datetime import datetime
from io import BytesIO
from .file_storage import store_Github

def read_xl(file):
    workbook = openpyxl.open(file)
    worksheet = workbook[workbook.sheetnames[0]]
    return worksheet


def validate(response):
    if response:
        return "gültig"
    else:
        return "nicht gültig"


def validate_UIDs(file, max_rows, recycle=[], iteration = 0, max_iter=1, wb=None):
    print(f"Iteration:  {iteration}")
    print(file)
    if iteration == 0:
        workbook = openpyxl.open(file)
    else:
        workbook = wb
    print(workbook)
    worksheet = workbook[workbook.sheetnames[0]]
    failures = []
    if not max_rows:
        max_rows = worksheet.max_row
    client = Client("http://ec.europa.eu/taxation_customs/vies/checkVatService.wsdl")
    for j, row in enumerate(worksheet.iter_rows(min_row=2, max_row=max_rows)):
        if (iteration > 0) & (iteration <= max_iter) & (j not in recycle):
            continue
        print(j)
        if j % 100 == 0:
            print(j)
        if row[0].value is not None:
            try:
                country = row[0].value[:2]
                vat = row[0].value[2:]
                response = client.service.checkVat(countryCode=country, vatNumber=vat)['valid']
                row[4].value = validate(response)
                with open('logfile.csv', 'a') as f:
                    f.write(str(j + 1) + ";" + row[0].value + ";" + row[4].value + '\n')

            except Exception as e:
                # Add Index of failure to list
                failures.append(j)
                row[4].value = "ungültiger input"
                with open('logfile.csv', 'a') as f:
                    f.write(str(j + 1) + ";" + row[0].value + ";" + row[4].value + ";" + str(e) + '\n')
        else:
            row[4].value = 'blank'

    # run the file again with failures
    if iteration+1<=max_iter:
        return validate_UIDs(file, max_rows, recycle=failures, iteration=iteration+1, max_iter=max_iter, wb=workbook)
    else:
        return workbook



def handle_uploaded_file(file, rows=None):

    filetype = str(file).split(".")[-1]
    if filetype not in ("xlsx", "xls"):
        print("Bitte ein Excel file verwenden")
    else:
        workbook = validate_UIDs(file, max_rows=None, max_iter=3)
        print(workbook)
        print("workbook was validated, next thing is saving")
        # store_Github(workbook)
        filename = str(file).split(".")[0] + '_checked.xlsx'
        workbook.save(f'media/validated_Documents/{filename}')
        print("The file was stored")
        return filename

